from __future__ import annotations

import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models import WeeklyPlan
from app.utils import week_in_month_for_period


def _progress_display(item) -> str:
    if item.progress_percent is not None:
        return f"{int(item.progress_percent)}%"
    if item.progress_text:
        return item.progress_text
    return "/"


def _details_text(item) -> str:
    if item.details:
        lines = []
        for idx, d in enumerate(sorted(item.details, key=lambda x: x.sort_no), start=1):
            lines.append(f"{idx}. {d.content}")
        return "\n".join(lines)
    if item.detail_text:
        return item.detail_text
    return ""


def _item_total_hours(item) -> float:
    if item.details:
        total = 0.0
        for d in item.details:
            if d.hours is None:
                continue
            total += float(d.hours)
        return total
    return float(item.estimated_hours or 0)


def export_plan_xlsx(plan: WeeklyPlan) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "周计划"

    columns = ["所属大类", "子项目", "本周目标", "进度", "目标细节", "预计工时", "工时合计"]
    col_widths = [18, 16, 16, 10, 80, 12, 12]
    for idx, w in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = w

    period = plan.period
    wim = week_in_month_for_period(period.start_date)
    title = f"周计划-[{plan.owner.name}]-[{period.year}]年[{period.month}月第{wim}周]"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=title)
    ws.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    start_row = 3
    total_est = 0.0
    total_sum = 0.0

    items = sorted(plan.items, key=lambda x: x.sort_no)
    for i, item in enumerate(items, start=0):
        row = start_row + i
        category = item.category.name if item.category else ""
        sub = item.sub_project.name if item.sub_project else ""
        goal = item.weekly_goal or ""
        progress = _progress_display(item)
        details = _details_text(item)
        est = float(item.estimated_hours) if item.estimated_hours is not None else 0.0
        item_sum = _item_total_hours(item)

        total_est += est
        total_sum += item_sum

        values = [category, sub, goal, progress, details, est, item_sum]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            if col in (1, 2, 3, 4, 6, 7):
                cell.alignment = center
            else:
                cell.alignment = left_wrap

        ws.row_dimensions[row].height = 48

    total_row = start_row + len(items)
    for col in range(1, len(columns) + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.border = border
        cell.alignment = center

    ws.cell(row=total_row, column=5, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_est).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_sum).font = Font(bold=True)

    ws.freeze_panes = "A3"
    ws.row_dimensions[2].height = 20
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
